import csv
import json
import logging
import os
from pathlib import Path
from openpyxl import Workbook, load_workbook
import pandas as pd
from src.logger import Logger

class DataHandler:
    def __init__(self):
        self.logger = Logger()
        # Constructor code here

    def verify_path(self, file_path: Path) -> bool:
        if file_path is None:
            print("Path is None")
            return False

        if file_path.exists():
            print("Path exists.")
            return True
        else:
            print("Path does not exist.")
            return False


    def write_to_file(self, file_path: Path, df, mode: str = 'w', encode: str = 'utf-8'):
        file_ext = file_path.suffix.lower()

        if file_ext == '.csv':
            if mode == 'a' and file_path.exists():
                df.to_csv(file_path, mode='a', header=False, index=False, encoding=encode)
            else:
                df.to_csv(file_path, mode='w', index=False, encoding=encode)
        elif file_ext == '.xlsx':
            if mode == 'a' and file_path.exists():
                with pd.ExcelWriter(file_path, mode='a', engine='openpyxl') as writer:
                    writer.book = load_workbook(file_path)
                    writer.sheets = dict((ws.title, ws) for ws in writer.book.worksheets)
                    df.to_excel(writer, sheet_name='Sheet1', index=False, header=False)
            else:
                with pd.ExcelWriter(file_path, engine='xlsxwriter') as writer:
                    df.to_excel(writer, sheet_name='Sheet1', index=False)
        else:
            raise ValueError("Unsupported file format. Only CSV and XLSX files are supported.")


    def load_file_to_pandas(self, file_path: Path, sheet_name: str = None) -> pd.DataFrame:
        if file_path is None:
            print("File path is None")
            return None

        if not file_path.exists():
            print("File does not exist")
            return None

        file_format = file_path.suffix.lower()[1:]  # Remove the dot (.) from the extension

        if file_format == 'csv':
            return self.__load_csv_to_pandas(file_path)
        elif file_format == 'xlsx':
            return self.__load_excel_to_pandas(file_path, sheet_name)
        else:
            print("Unsupported file format")
            return None

    def __load_csv_to_pandas(self, file_path: Path) -> pd.DataFrame:
        if not self.verify_path(file_path):
            return None

        try:
            df = pd.read_csv(file_path)
            print("CSV file loaded into pandas DataFrame successfully.")
            return df
        except Exception as e:
            print(f"Error occurred during loading CSV to pandas: {e}")

    def __load_excel_to_pandas(self, file_path: Path, sheet_name: str = None) -> pd.DataFrame:
        if not self.verify_path(file_path):
            return None

        try:
            workbook = load_workbook(file_path)
            if sheet_name:
                sheet = workbook[sheet_name]
            else:
                sheet = workbook.active
            data = sheet.values
            headers = next(data)
            df = pd.DataFrame(data, columns=headers)
            print("Excel file loaded into pandas DataFrame successfully.")
            return df
        except Exception as e:
            print(f"Error occurred during loading Excel to pandas: {e}")

    def save_data_params(self, key, data, file_path=None, file_name=None):
        # Set default file path if not provided
        if file_path is None:
            file_path = Path("../data/")

        # Create directory if it doesn't exist
        os.makedirs(file_path, exist_ok=True)

        # Create file path with key as the file name
        file_path = file_path / f"{file_name if file_name else 'data_params'}.json"

        # Prepare data to be saved
        data_to_save = {key: data}

        try:
            # Save data to JSON file
            with open(file_path, "w", encoding="utf-8") as file:
                json.dump(data_to_save, file, indent=4, ensure_ascii=False)
            # Log success message
            self.logger.log(f"Data saved successfully for key: {key}", self.logger.DEBUG)

        except Exception as e:
            # Log failure message
            self.logger.log(f"Failed to save data for key: {key}. Error: {e}", self.logger.DEBUG)

    def load_data_params(self, key, file_path=None, file_name=None):
        # Set default file path if not provided
        if file_path is None:
            file_path = Path("../data/")

        # Create file path with key as the file name
        file_path = file_path / f"{file_name if file_name else 'data_params'}.json"

        try:
            # Check if file exists
            if file_path.exists():
                # Read data from file
                with open(file_path, "r") as file:
                    data = json.load(file)

                # Check if key exists in the loaded data
                if key in data:
                    # Return the value associated with the key
                    return data[key]

            # Log message if key or file not found
            self.logger.log(f"Data not found for key: {key} or file not found: {file_path}", self.logger.DEBUG)

        except Exception as e:
            # Log failure message
            self.logger.log(f"Failed to load data for key: {key}. Error: {e}", self.logger.DEBUG)

        # Return None if data not found or an error occurred
        return None

